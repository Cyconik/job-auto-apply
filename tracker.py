"""
tracker.py — Application log, load, export
"""
import json
from pathlib import Path
from datetime import datetime

LOG = Path("job_applications.json")


def log(app: dict):
    all_apps = load_all()
    existing = {a.get("url") for a in all_apps}
    if app.get("url") in existing:
        return
    all_apps.append(app)
    LOG.write_text(json.dumps(all_apps, indent=2, ensure_ascii=False), encoding="utf-8")


def load_all() -> list:
    if not LOG.exists():
        return []
    try:
        return json.loads(LOG.read_text(encoding="utf-8"))
    except Exception:
        return []


def clear_all():
    if LOG.exists():
        LOG.unlink()


def export_excel(apps: list) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"

        hfill = PatternFill("solid", fgColor="667EEA")
        hfont = Font(color="FFFFFF", bold=True)
        thin  = Border(*[Side(style="thin")] * 0,
                       left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"),  bottom=Side(style="thin"))

        sfill = {"applied": PatternFill("solid", fgColor="C8E6C9"),
                 "failed":  PatternFill("solid", fgColor="FFCDD2"),
                 "pending": PatternFill("solid", fgColor="FFF9C4")}

        cols = [
            ("Date/Time",  "timestamp", 20),
            ("Company",    "company",   22),
            ("Job Title",  "job_title", 28),
            ("Platform",   "platform",  18),
            ("Location",   "location",  20),
            ("Salary",     "salary",    18),
            ("Status",     "status",    12),
            ("Note",       "note",      40),
            ("URL",        "url",       45),
        ]

        ws.row_dimensions[1].height = 28
        for ci, (header, _, width) in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=header)
            c.font      = hfont
            c.fill      = hfill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin
            ws.column_dimensions[get_column_letter(ci)].width = width

        for ri, app in enumerate(apps, 2):
            ws.row_dimensions[ri].height = 18
            fill = sfill.get(app.get("status",""))
            for ci, (_, key, _) in enumerate(cols, 1):
                c = ws.cell(row=ri, column=ci, value=str(app.get(key,"") or ""))
                c.border    = thin
                c.alignment = Alignment(vertical="center")
                if fill:
                    c.fill = fill

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        total   = len(apps)
        applied = sum(1 for a in apps if a.get("status")=="applied")
        failed  = sum(1 for a in apps if a.get("status")=="failed")
        pending = sum(1 for a in apps if a.get("status")=="pending")
        for r, (lbl, val) in enumerate([
            ("Total Applications", total),
            ("Applied ✅",         applied),
            ("Failed ❌",          failed),
            ("Pending ⚠️",         pending),
            ("Success Rate",       f"{int(applied/total*100) if total else 0}%"),
        ], 1):
            ws2.cell(row=r, column=1, value=lbl)
            ws2.cell(row=r, column=2, value=val)
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 12

        out = "job_applications.xlsx"
        wb.save(out)
        return out
    except ImportError:
        # CSV fallback
        import csv
        out = "job_applications.csv"
        if apps:
            with open(out, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=list(apps[0].keys()))
                w.writeheader()
                w.writerows(apps)
        return out
